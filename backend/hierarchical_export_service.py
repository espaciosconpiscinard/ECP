"""
Servicio de Exportación Jerárquica - Plantillas Separadas por Entidad
Cada entidad tiene su propia plantilla con validaciones y dropdowns
"""

import pandas as pd
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import datetime
from typing import List, Dict

# Colores
COLOR_HEADER = "4472C4"
COLOR_REQUIRED = "FFF2CC"
COLOR_OPTIONAL = "E7E6E6"

def _apply_header_style(ws, row_num=1):
    """Aplica estilo al header"""
    for cell in ws[row_num]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

def _apply_required_fill(ws, col_letter, start_row=2, end_row=100):
    """Aplica color amarillo a columnas obligatorias"""
    for row in range(start_row, end_row + 1):
        ws[f"{col_letter}{row}"].fill = PatternFill(start_color=COLOR_REQUIRED, end_color=COLOR_REQUIRED, fill_type="solid")

def _apply_optional_fill(ws, col_letter, start_row=2, end_row=100):
    """Aplica color gris a columnas opcionales"""
    for row in range(start_row, end_row + 1):
        ws[f"{col_letter}{row}"].fill = PatternFill(start_color=COLOR_OPTIONAL, end_color=COLOR_OPTIONAL, fill_type="solid")

def _add_dropdown(ws, col_letter, options, start_row=2, end_row=100):
    """Agrega lista desplegable a una columna"""
    dv = DataValidation(type="list", formula1=f'"{",".join(options)}"', allow_blank=True)
    dv.error = 'Por favor selecciona un valor de la lista'
    dv.errorTitle = 'Valor Inválido'
    ws.add_data_validation(dv)
    dv.add(f'{col_letter}{start_row}:{col_letter}{end_row}')

# ========== PASO 1: PLANTILLA DE CLIENTES ==========
def generate_customers_template() -> BytesIO:
    """Genera plantilla Excel para importar clientes"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Clientes"
    
    # Headers
    headers = ["Nombre *", "Email", "Teléfono", "Cédula/RNC", "Dirección"]
    ws.append(headers)
    _apply_header_style(ws)
    
    # Marcar columnas obligatorias y opcionales
    _apply_required_fill(ws, "A")  # Nombre
    _apply_optional_fill(ws, "B")  # Email
    _apply_optional_fill(ws, "C")  # Teléfono
    _apply_optional_fill(ws, "D")  # Cédula
    _apply_optional_fill(ws, "E")  # Dirección
    
    # Ejemplo
    ws.append(["Juan Pérez", "juan@email.com", "809-555-1234", "001-1234567-8", "Calle Principal #123"])
    
    # Ajustar anchos
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 40
    
    # Guardar en BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output

# ========== PASO 2: PLANTILLA DE CATEGORÍAS DE VILLAS ==========
def generate_villa_categories_template() -> BytesIO:
    """Genera plantilla Excel para importar categorías de villas"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Categorías"
    
    # Headers
    headers = ["Nombre Categoría *", "Descripción"]
    ws.append(headers)
    _apply_header_style(ws)
    
    # Marcar columnas
    _apply_required_fill(ws, "A")  # Nombre
    _apply_optional_fill(ws, "B")  # Descripción
    
    # Ejemplos
    ws.append(["VIP", "Villas con características premium"])
    ws.append(["Estándar", "Villas básicas"])
    ws.append(["Familiar", "Villas grandes para familias"])
    
    # Ajustar anchos
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 50
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output

# ========== PASO 3: PLANTILLA DE VILLAS ==========
async def generate_villas_template(db) -> BytesIO:
    """Genera plantilla Excel para importar villas con dropdown de categorías"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Villas"
    
    # Headers
    headers = ["Código Villa *", "Nombre Villa *", "Categoría", "Precio Cliente", "Precio Propietario", 
               "Descripción", "Horario Check-in", "Horario Check-out"]
    ws.append(headers)
    _apply_header_style(ws)
    
    # Obtener categorías existentes para dropdown
    categories = await db.categories.find({'is_active': True}, {"_id": 0, "name": 1}).to_list(100)
    category_names = [cat["name"] for cat in categories] if categories else ["VIP", "Estándar", "Familiar"]
    
    # Crear hoja oculta para las categorías (solución para evitar límite de 255 caracteres)
    ws_categories = wb.create_sheet("_Categorias")
    ws_categories.sheet_state = 'hidden'
    
    # Escribir categorías en hoja oculta
    for idx, cat_name in enumerate(category_names, start=1):
        ws_categories[f'A{idx}'] = cat_name
    
    # Marcar columnas
    _apply_required_fill(ws, "A")  # Código (obligatorio)
    _apply_required_fill(ws, "B")  # Nombre (obligatorio)
    _apply_optional_fill(ws, "C")  # Categoría (opcional)
    _apply_optional_fill(ws, "D")  # Precio Cliente (opcional - para precios flexibles)
    _apply_optional_fill(ws, "E")  # Precio Propietario (opcional)
    _apply_optional_fill(ws, "F")  # Descripción
    _apply_optional_fill(ws, "G")  # Check-in
    _apply_optional_fill(ws, "H")  # Check-out
    
    # Dropdown para categorías usando referencia a hoja oculta
    if category_names:
        dv = DataValidation(
            type="list", 
            formula1=f'=_Categorias!$A$1:$A${len(category_names)}',
            allow_blank=True
        )
        dv.error = 'Por favor selecciona una categoría de la lista'
        dv.errorTitle = 'Categoría Inválida'
        ws.add_data_validation(dv)
        dv.add('C2:C100')
    
    # Ejemplo
    ws.append(["ECPVSH", "Villa Shangrila", category_names[0] if category_names else "VIP", 
               "15000", "7000", "Villa con piscina grande", "08:00", "18:00"])
    
    # Ajustar anchos
    for col, width in [("A", 15), ("B", 25), ("C", 15), ("D", 18), ("E", 20), ("F", 40), ("G", 15), ("H", 15)]:
        ws.column_dimensions[col].width = width
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output

# ========== PASO 4: PLANTILLA DE SERVICIOS EXTRA ==========
def generate_services_template() -> BytesIO:
    """Genera plantilla Excel para importar servicios extra"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Servicios Extra"
    
    # Headers
    headers = ["Nombre Servicio *", "Precio *", "Descripción"]
    ws.append(headers)
    _apply_header_style(ws)
    
    # Marcar columnas
    _apply_required_fill(ws, "A")  # Nombre
    _apply_required_fill(ws, "B")  # Precio
    _apply_optional_fill(ws, "C")  # Descripción
    
    # Ejemplos
    ws.append(["Chef privado", "5000", "Chef para cocinar en la villa"])
    ws.append(["DJ", "8000", "DJ para eventos"])
    ws.append(["Decoración", "3000", "Decoración temática"])
    
    # Ajustar anchos
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 50
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output

# ========== PASO 5: PLANTILLA DE CATEGORÍAS DE GASTOS ==========
def generate_expense_categories_template() -> BytesIO:
    """Genera plantilla Excel para importar categorías de gastos"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Categorías Gastos"
    
    # Headers
    headers = ["Nombre Categoría *", "Descripción"]
    ws.append(headers)
    _apply_header_style(ws)
    
    # Marcar columnas
    _apply_required_fill(ws, "A")  # Nombre
    _apply_optional_fill(ws, "B")  # Descripción
    
    # Ejemplos
    ws.append(["Mantenimiento", "Gastos de mantenimiento general"])
    ws.append(["Servicios Públicos", "Agua, luz, internet"])
    ws.append(["Personal", "Pagos a empleados"])
    
    # Ajustar anchos
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 50
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output

# ========== PASO 6: PLANTILLA DE RESERVACIONES ==========
async def generate_reservations_template(db) -> BytesIO:
    """Genera plantilla Excel para importar reservaciones con dropdowns"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Reservaciones"
    
    # Headers
    headers = ["N° Factura *", "Fecha Reservación *", "Cliente *", "Villa *", "Tipo Alquiler *", 
               "Check-in", "Check-out", "N° Personas", "Precio *", "Moneda *", "Método Pago *", 
               "Monto Pagado", "Depósito", "Incluir ITBIS", "Servicios Extra", "Notas"]
    ws.append(headers)
    _apply_header_style(ws)
    
    # Obtener datos para dropdowns
    customers = await db.customers.find({}, {"_id": 0, "name": 1}).to_list(500)
    customer_names = [c["name"] for c in customers] if customers else ["Ejemplo Cliente"]
    
    villas = await db.villas.find({}, {"_id": 0, "code": 1}).to_list(100)
    villa_codes = [v["code"] for v in villas] if villas else ["ECPVSH"]
    
    services = await db.extra_services.find({}, {"_id": 0, "name": 1}).to_list(100)
    service_names = [s["name"] for s in services] if services else ["Chef privado"]
    
    # Opciones predefinidas
    rental_types = ["pasadia", "amanecida", "evento"]
    currencies = ["DOP", "USD"]
    payment_methods = ["efectivo", "transferencia", "deposito", "mixto"]
    yes_no = ["SI", "NO"]
    
    # Marcar columnas obligatorias y opcionales
    _apply_required_fill(ws, "A")  # Factura
    _apply_required_fill(ws, "B")  # Fecha
    _apply_required_fill(ws, "C")  # Cliente
    _apply_required_fill(ws, "D")  # Villa
    _apply_required_fill(ws, "E")  # Tipo
    _apply_optional_fill(ws, "F")  # Check-in
    _apply_optional_fill(ws, "G")  # Check-out
    _apply_optional_fill(ws, "H")  # Personas
    _apply_required_fill(ws, "I")  # Precio
    _apply_required_fill(ws, "J")  # Moneda
    _apply_required_fill(ws, "K")  # Método pago
    _apply_optional_fill(ws, "L")  # Monto pagado
    _apply_optional_fill(ws, "M")  # Depósito
    _apply_optional_fill(ws, "N")  # ITBIS
    _apply_optional_fill(ws, "O")  # Servicios
    _apply_optional_fill(ws, "P")  # Notas
    
    # Agregar dropdowns
    _add_dropdown(ws, "C", customer_names[:50])  # Limitar a 50 para evitar problemas
    _add_dropdown(ws, "D", villa_codes)
    _add_dropdown(ws, "E", rental_types)
    _add_dropdown(ws, "J", currencies)
    _add_dropdown(ws, "K", payment_methods)
    _add_dropdown(ws, "N", yes_no)
    
    # Ejemplo
    ws.append(["1001", "15/10/2025", customer_names[0] if customer_names else "Juan Pérez", 
               villa_codes[0] if villa_codes else "ECPVSH", "pasadia", "08:00", "18:00", "10", 
               "15000", "DOP", "transferencia", "10000", "2000", "SI", "", "Reserva confirmada"])
    
    # Ajustar anchos
    for col, width in [("A", 12), ("B", 15), ("C", 25), ("D", 12), ("E", 12), ("F", 10), 
                       ("G", 10), ("H", 10), ("I", 12), ("J", 10), ("K", 15), ("L", 15), 
                       ("M", 12), ("N", 12), ("O", 30), ("P", 40)]:
        ws.column_dimensions[col].width = width
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output

# ========== FUNCIONES AUXILIARES ==========
async def get_all_templates_info(db) -> List[Dict]:
    """Retorna información de todas las plantillas en orden jerárquico"""
    
    # Contar registros existentes
    customers_count = await db.customers.count_documents({})
    categories_count = await db.categories.count_documents({})
    villas_count = await db.villas.count_documents({})
    services_count = await db.extra_services.count_documents({})
    expense_categories_count = await db.expense_categories.count_documents({})
    reservations_count = await db.reservations.count_documents({})
    
    return [
        {
            "step": 1,
            "name": "Clientes",
            "key": "customers",
            "description": "Importa la base de clientes",
            "required": True,
            "count": customers_count,
            "icon": "👥"
        },
        {
            "step": 2,
            "name": "Categorías de Villas",
            "key": "villa_categories",
            "description": "Importa categorías (VIP, Estándar, etc.)",
            "required": False,
            "count": categories_count,
            "icon": "🏷️"
        },
        {
            "step": 3,
            "name": "Villas",
            "key": "villas",
            "description": "Importa villas con sus precios",
            "required": True,
            "count": villas_count,
            "icon": "🏠"
        },
        {
            "step": 4,
            "name": "Servicios Extra",
            "key": "services",
            "description": "Importa servicios adicionales (DJ, Chef, etc.)",
            "required": False,
            "count": services_count,
            "icon": "✨"
        },
        {
            "step": 5,
            "name": "Categorías de Gastos",
            "key": "expense_categories",
            "description": "Importa categorías de gastos",
            "required": False,
            "count": expense_categories_count,
            "icon": "📂"
        },
        {
            "step": 6,
            "name": "Reservaciones",
            "key": "reservations",
            "description": "Importa facturas históricas con dropdowns",
            "required": True,
            "count": reservations_count,
            "icon": "📅",
            "depends_on": ["customers", "villas"]
        }
    ]
